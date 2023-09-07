import customtkinter as ctk
import tkinter
import openpyxl


class BancoApp:
    def __init__(self, root):
        self.root = root
        self.root.title('SGBANK')
        self.root.geometry('400x300')
        self.root.minsize(500, 400)
        self.root.maxsize(500, 400)

        self.initialize_ui()

    def initialize_ui(self):
        # Elementos da interface principal
        self.margin(30)
        title = ctk.CTkLabel(self.root, text='Bem-vindo ao SGBANK', text_color='#fff', font=('Montserrat', 26, 'bold'))
        title.pack()

        self.margin(5)
        subtitle = ctk.CTkLabel(self.root, text='O que deseja fazer?', text_color='#fff',
                                font=('Montserrat', 20, 'bold'))
        subtitle.pack()

        self.margin(20)
        extrato = ctk.CTkButton(self.root, text='Extrato', text_color='#fff', font=('Montserrat', 20, 'bold'),
                                command=self.janela_extrato)
        extrato.pack()

        self.margin(20)
        depositar = ctk.CTkButton(self.root, text='Depositar', text_color='#fff', font=('Montserrat', 20, 'bold'),
                                  command=self.janela_depositar)
        depositar.pack()

        self.margin(20)
        sacar = ctk.CTkButton(self.root, text='Sacar', text_color='#fff', font=('Montserrat', 20, 'bold'),
                              command=self.janela_sacar)
        sacar.pack()

        self.margin(20)
        transferir = ctk.CTkButton(self.root, text='Transferir', text_color='#fff', font=('Montserrat', 20, 'bold'),
                                   command=self.janela_transferir)
        transferir.pack()

        self.margin(20)
        create_user = ctk.CTkButton(self.root, text='Criar conta', text_color='#fff', font=('Montserrat', 20, 'bold'),
                                    command=self.janela_user)
        create_user.pack()

    def margin(self, height):
        margem = ctk.CTkCanvas(self.root, highlightthickness=0, height=height, background='#242424')
        margem.pack()

    def janela_user(self):
        criar_conta_app = CriarContaApp(self)
        self.root.wait_window(criar_conta_app.ja_user)

    def janela_extrato(self):
        extrato_app = ExtratoApp(self)
        self.root.wait_window(extrato_app.ja_extrato)

    def janela_depositar(self):
        depositar_app = DepositarApp(self)
        self.root.wait_window(depositar_app.ja_depo)

    def janela_transferir(self):
        transferir_app = TransferirApp(self)
        self.root.wait_window(transferir_app.ja_transferir)

    def janela_sacar(self):
        sacar_app = SacarApp(self)
        self.root.wait_window(sacar_app.ja_sacar)


class CriarContaApp:
    def __init__(self, parent):
        self.parent = parent
        self.ja_user = ctk.CTkToplevel(parent.root)
        self.ja_user.geometry('400x400')
        self.ja_user.title('Criar Conta')
        self.ja_user.maxsize(500, 400)
        self.ja_user.minsize(500, 400)
        self.ja_user.focus_force()
        self.ja_user.grab_set()

        self.initialize_ui()
        self.conta = 1000

    def initialize_ui(self):
        titulo = ctk.CTkLabel(self.ja_user, text='Criar conta', text_color='#fff',
                              font=('Montserrat', 26, 'bold'))
        titulo.pack(pady=30)

        label_nome = ctk.CTkLabel(self.ja_user, text='Informe seu nome:', text_color='#fff',
                                  font=('Montserrat', 18, 'bold'))
        label_nome.place(x=30, y=90)

        self.entry_nome = ctk.CTkEntry(self.ja_user, placeholder_text='Insira seu nome', placeholder_text_color='#fff',
                                       font=('Montserrat', 12, 'bold'), fg_color='transparent')
        self.entry_nome.place(x=200, y=90)

        label_sobrenome = ctk.CTkLabel(self.ja_user, text='Informe seu sobrenome:', text_color='#fff',
                                       font=('Montserrat', 18, 'bold'))
        label_sobrenome.place(x=30, y=130)

        self.entry_sobrenome = ctk.CTkEntry(self.ja_user, placeholder_text='Insira seu sobrenome', placeholder_text_color='#fff',
                                            font=('Montserrat', 12, 'bold'), fg_color='transparent')
        self.entry_sobrenome.place(x=250, y=130)

        label_nascimento = ctk.CTkLabel(self.ja_user, text='Informe sua data de nascimento:', text_color='#fff',
                                        font=('Montserrat', 18, 'bold'))
        label_nascimento.place(x=30, y=170)

        self.entry_nascimento = ctk.CTkEntry(self.ja_user, placeholder_text=' 00/00/0000 ', placeholder_text_color='#fff',
                                             font=('Montserrat', 12, 'bold'), fg_color='transparent')
        self.entry_nascimento.place(x=320, y=170)

        label_cpf = ctk.CTkLabel(self.ja_user, text='Informe seu CPF:', text_color='#fff',
                                 font=('Montserrat', 18, 'bold'))
        label_cpf.place(x=30, y=210)

        self.entry_cpf = ctk.CTkEntry(self.ja_user, placeholder_text=' 000.000.000-00 ', placeholder_text_color='#fff',
                                      font=('Montserrat', 12, 'bold'), fg_color='transparent')
        self.entry_cpf.place(x=185, y=210)

        label_senha = ctk.CTkLabel(self.ja_user, text='Crie sua senha(com 4 digitos): ', text_color='#fff',
                                   font=('Montserrat', 18, 'bold'))
        label_senha.place(x=30, y=250)

        self.entry_senha = ctk.CTkEntry(self.ja_user, placeholder_text=' **** ', placeholder_text_color='#fff',
                                        font=('Montserrat', 12, 'bold'), fg_color='transparent', show='*')
        self.entry_senha.place(x=300, y=250)

        button_criar = ctk.CTkButton(self.ja_user, text='Criar Conta', text_color='#fff', font=('Montserrat', 18, 'bold'),
                                     command=self.criar_conta)
        button_criar.place(x=175, y=300)

    def criar_conta(self):
        nome = self.entry_nome.get()
        sobrenome = self.entry_sobrenome.get()
        data = self.entry_nascimento.get()
        cpf = self.entry_cpf.get()
        senha = self.entry_senha.get()
        saldo = 100

        fichario = openpyxl.load_workbook('cadastros.xlsx')
        folha = fichario.active

        if nome == "" or sobrenome == "" or data == "" or cpf == "" or senha == "":
            tkinter.messagebox.showwarning('Sistema', "ERROR\nPor favor prenchar todos os campos!!")
            self.entry_nome.delete(0, tkinter.END)
            self.entry_sobrenome.delete(0, tkinter.END)
            self.entry_nascimento.delete(0, tkinter.END)
            self.entry_cpf.delete(0, tkinter.END)
            self.entry_senha.delete(0, tkinter.END)
        else:
            folha.cell(column=1, row=folha.max_row + 1, value=self.conta)
            folha.cell(column=2, row=folha.max_row, value=nome)
            folha.cell(column=3, row=folha.max_row, value=sobrenome)
            folha.cell(column=4, row=folha.max_row, value=data)
            folha.cell(column=5, row=folha.max_row, value=cpf)
            folha.cell(column=6, row=folha.max_row, value=senha)
            folha.cell(column=7, row=folha.max_row, value=saldo)
            fichario.save(r'cadastros.xlsx')
            tkinter.messagebox.showinfo('Sistema', f"Conta criada com sucesso!! \nO numero da sua conta é {self.conta}"
                                                   f"\n Você ganhou um saldo de R$: 100,00 por cadastrar!!")
            self.entry_nome.delete(0, tkinter.END)
            self.entry_sobrenome.delete(0, tkinter.END)
            self.entry_nascimento.delete(0, tkinter.END)
            self.entry_cpf.delete(0, tkinter.END)
            self.entry_senha.delete(0, tkinter.END)
            self.conta += 1


class ExtratoApp:
    def __init__(self, parent):
        self.parent = parent
        self.ja_extrato = ctk.CTkToplevel(parent.root)
        self.ja_extrato.geometry('400x300')
        self.ja_extrato.title('Extrato')
        self.ja_extrato.maxsize(500, 300)
        self.ja_extrato.minsize(500, 300)
        self.ja_extrato.focus_force()
        self.ja_extrato.grab_set()

        self.initialize_ui()

    def initialize_ui(self):
        titulo = ctk.CTkLabel(self.ja_extrato, text='Extrato', text_color='#fff',
                              font=('Montserrat', 26, 'bold'))
        titulo.pack(pady=30)

        label_conta = ctk.CTkLabel(self.ja_extrato, text='Informe o numero da sua conta: ', text_color='#fff',
                                   font=('Montserrat', 17, 'bold'))
        label_conta.place(x=30, y=90)

        self.entry_conta = ctk.CTkEntry(self.ja_extrato, placeholder_text='Insira a conta', placeholder_text_color='#fff',
                                   font=('Montserrat', 12, 'bold'), fg_color='transparent')
        self.entry_conta.place(x=290, y=90)

        label_senha = ctk.CTkLabel(self.ja_extrato, text='Informe a senha: ', text_color='#fff',
                                   font=('Montserrat', 17, 'bold'))
        label_senha.place(x=30, y=130)

        self.entry_senha = ctk.CTkEntry(self.ja_extrato, placeholder_text='Insira a senha', placeholder_text_color='#fff',
                                        font=('Montserrat', 12, 'bold'), fg_color='transparent')
        self.entry_senha.place(x=170, y=130)

        extrato_button = ctk.CTkButton(self.ja_extrato, text='Verificar extrato', text_color='#fff',
                                       font=('Montserrat', 18, 'bold'), width=30, command=self.extrato)
        extrato_button.place(x=170, y=180)

        self.label_extrato = ctk.CTkLabel(self.ja_extrato, text='', text_color='#fff',
                                          font=('Montserrat', 24, 'bold'))
        self.label_extrato.place(x=100, y=240)

    def extrato(self):
        fichario = openpyxl.load_workbook('cadastros.xlsx')
        folha = fichario.active
        dados = []
        e_conta = self.entry_conta.get()
        e_senha = self.entry_senha.get()

        for row in folha.iter_cols(values_only=True):
            dados.append(list(row))

        dicionario = {}

        for i in range(len(dados[0])):
            chave = dados[0][i]
            valores = [dados[j][i] for j in range(1, len(dados))]  # Valores associados à chave
            dicionario[chave] = valores

        if e_conta == "" or e_senha == "":
            self.label_extrato.configure(text='')
            tkinter.messagebox.showwarning('Sistema', "ERROR\nPor favor prenchar todos os campos!!")
        else:
            if int(e_conta) in dicionario:
                if dicionario[int(e_conta)][4] == str(e_senha):
                    self.label_extrato.configure(text=f'Seu saldo é de R${dicionario[int(e_conta)][5]:.2f}')
                else:
                    self.label_extrato.configure(text='')
                    tkinter.messagebox.showwarning('Sistema', "ERROR\nSenha invalida!!")
            else:
                self.label_extrato.configure(text='')
                tkinter.messagebox.showwarning('Sistema', "ERROR\nConta invalida!!")


class DepositarApp:
    def __init__(self, parent):
        self.parent = parent
        self.ja_depo = ctk.CTkToplevel(parent.root)
        self.ja_depo.geometry('500x250')
        self.ja_depo.title('Depositar')
        self.ja_depo.maxsize(500, 250)
        self.ja_depo.minsize(500, 250)
        self.ja_depo.focus_force()
        self.ja_depo.grab_set()

        self.initialize_ui()

    def initialize_ui(self):
        titulo = ctk.CTkLabel(self.ja_depo, text='Depositar', text_color='#fff',
                              font=('Montserrat', 26, 'bold'))
        titulo.pack(pady=30)

        label_conta = ctk.CTkLabel(self.ja_depo, text='Informe o numero da sua conta: ', text_color='#fff',
                                   font=('Montserrat', 17, 'bold'))
        label_conta.place(x=30, y=90)

        self.entry_conta = ctk.CTkEntry(self.ja_depo, placeholder_text='Insira a conta', placeholder_text_color='#fff',
                                   font=('Montserrat', 12, 'bold'), fg_color='transparent')
        self.entry_conta.place(x=290, y=90)

        label_valor = ctk.CTkLabel(self.ja_depo, text='Informe o valor do deposito: ', text_color='#fff',
                                   font=('Montserrat', 17, 'bold'))
        label_valor.place(x=30, y=130)

        self.entry_valor = ctk.CTkEntry(self.ja_depo, placeholder_text='Insira o valor', placeholder_text_color='#fff',
                                   font=('Montserrat', 12, 'bold'), fg_color='transparent')
        self.entry_valor.place(x=260, y=130)

        deposito_button = ctk.CTkButton(self.ja_depo, text='Fazer deposito', text_color='#fff',
                                        font=('Montserrat', 18, 'bold'),
                                        width=30, command=self.deposito)
        deposito_button.place(x=170, y=180)

    def deposito(self):
        fichario = openpyxl.load_workbook('cadastros.xlsx')
        folha = fichario.active
        dados = []
        e_conta = self.entry_conta.get()
        e_valor = self.entry_valor.get()

        for row in folha.iter_cols(values_only=True):
            dados.append(list(row))

        dicionario = {}

        for i in range(len(dados)):
            chave = dados[0][i]
            valores = [dados[j][i] for j in range(1, len(dados))]
            dicionario[chave] = valores

        if e_conta == "" or e_valor == "":
            tkinter.messagebox.showwarning('Sistema', "ERROR\nPor favor prenchar todos os campos!!")
        else:
            if int(e_conta) in dicionario:
                if float(e_valor) > 0:
                    for row in folha.iter_cols(min_row=2, max_row=folha.max_row, min_col=1, max_col=1):
                        for cell in row:
                            if cell.value == int(e_conta):
                                valor = cell.row
                                folha[f'G{valor}'] = dicionario[int(e_conta)][5] + float(e_valor)
                    tkinter.messagebox.showwarning('Sistema', "Deposito realizado com sucesso!!")
                    self.entry_conta.delete(0, tkinter.END)
                    self.entry_valor.delete(0, tkinter.END)
                    fichario.save(r'cadastros.xlsx')
                else:
                    tkinter.messagebox.showwarning('Sistema', "ERROR\nValor invalido!!")
            else:
                tkinter.messagebox.showwarning('Sistema', "ERROR\nConta invalida!!")


class TransferirApp:
    def __init__(self, parent):
        self.parent = parent
        self.ja_transferir = ctk.CTkToplevel(parent.root)
        self.ja_transferir.geometry('500x350')
        self.ja_transferir.title('Transferir')
        self.ja_transferir.maxsize(500, 350)
        self.ja_transferir.minsize(500, 350)
        self.ja_transferir.focus_force()
        self.ja_transferir.grab_set()

        self.initialize_ui()

    def initialize_ui(self):
        titulo = ctk.CTkLabel(self.ja_transferir, text='Transferir', text_color='#fff',
                              font=('Montserrat', 26, 'bold'))
        titulo.pack(pady=30)

        label_conta = ctk.CTkLabel(self.ja_transferir, text='Informe o numero da sua conta: ', text_color='#fff',
                                   font=('Montserrat', 17, 'bold'))
        label_conta.place(x=30, y=90)

        self.entry_conta = ctk.CTkEntry(self.ja_transferir, placeholder_text='Insira a conta', placeholder_text_color='#fff',
                                   font=('Montserrat', 12, 'bold'), fg_color='transparent')
        self.entry_conta.place(x=290, y=90)

        label_senha = ctk.CTkLabel(self.ja_transferir, text='Informe a senha: ', text_color='#fff',
                                   font=('Montserrat', 17, 'bold'))
        label_senha.place(x=30, y=130)

        self.entry_senha = ctk.CTkEntry(self.ja_transferir, placeholder_text='Insira a senha', placeholder_text_color='#fff',
                                   font=('Montserrat', 12, 'bold'), fg_color='transparent')
        self.entry_senha.place(x=170, y=130)

        label_conta2 = ctk.CTkLabel(self.ja_transferir, text='Informe o numero da outra conta: ', text_color='#fff',
                                    font=('Montserrat', 17, 'bold'))
        label_conta2.place(x=30, y=170)

        self.entry_conta2 = ctk.CTkEntry(self.ja_transferir, placeholder_text='Insira a outra conta',
                                    placeholder_text_color='#fff',
                                    font=('Montserrat', 12, 'bold'), fg_color='transparent')

        self.entry_conta2.place(x=300, y=170)

        label_valor = ctk.CTkLabel(self.ja_transferir, text='Informe o valor da transferência: ', text_color='#fff',
                                   font=('Montserrat', 17, 'bold'))
        label_valor.place(x=30, y=210)

        self.entry_valor = ctk.CTkEntry(self.ja_transferir, placeholder_text='Insira o valor', placeholder_text_color='#fff',
                                   font=('Montserrat', 12, 'bold'), fg_color='transparent')
        self.entry_valor.place(x=290, y=210)

        transferir_button = ctk.CTkButton(self.ja_transferir, text='Fazer transferência', text_color='#fff',
                                          font=('Montserrat', 18, 'bold'), command=self.transferir)
        transferir_button.place(x=160, y=260)

    def transferir(self):
        fichario = openpyxl.load_workbook('cadastros.xlsx')
        folha = fichario.active
        dados = []
        e_conta = self.entry_conta.get()
        e_conta2 = self.entry_conta2.get()
        e_senha = self.entry_senha.get()
        e_valor = self.entry_valor.get()

        for row in folha.iter_cols(values_only=True):
            dados.append(list(row))

        dicionario = {}

        for i in range(len(dados)):
            chave = dados[0][i]
            valores = [dados[j][i] for j in range(1, len(dados))]
            dicionario[chave] = valores

        if e_conta == "" or e_conta2 == "" or e_valor == "" or e_senha == "":
            tkinter.messagebox.showwarning('Sistema', "ERROR\nPor favor prenchar todos os campos!!")
        else:
            if int(e_conta) in dicionario:
                if int(e_conta2) in dicionario:
                    if int(e_conta) != int(e_conta2):
                        if dicionario[int(e_conta)][4] == str(e_senha):
                            if float(e_valor) != 0:
                                if float(e_valor) <= dicionario[int(e_conta)][5]:
                                    for row in folha.iter_cols(min_row=2, max_row=folha.max_row, min_col=1, max_col=1):
                                        for cell in row:
                                            if cell.value == int(e_conta2):
                                                valor = cell.row
                                                folha[f'G{valor}'] = dicionario[int(e_conta2)][5] + float(e_valor)
                                            if cell.value == int(e_conta):
                                                valor = cell.row
                                                folha[f'G{valor}'] = dicionario[int(e_conta)][5] - float(e_valor)
                                    tkinter.messagebox.showwarning('Sistema', "Tranferencia realizada com sucesso!!")
                                    fichario.save(r'cadastros.xlsx')
                                    self.entry_conta.delete(0, tkinter.END)
                                    self.entry_senha.delete(0, tkinter.END)
                                    self.entry_valor.delete(0, tkinter.END)
                                else:
                                    tkinter.messagebox.showwarning('Sistema',
                                                                   "ERROR\nVocê não tem saldo para realizar a transferencia!!")
                            else:
                                tkinter.messagebox.showwarning('Sistema', "ERROR\nValor invalido!!")
                        else:
                            tkinter.messagebox.showwarning('Sistema', "ERROR\nSenha invalida!!")
                    else:
                        tkinter.messagebox.showwarning('Sistema',
                                                       "ERROR\nVocê não poder transferir para sua propria conta!!")
                else:
                    tkinter.messagebox.showwarning('Sistema', "ERROR\nConta que deseja transferir estar invalida!!")
            else:
                tkinter.messagebox.showwarning('Sistema', "ERROR\nConta invalida!!")


class SacarApp:
    def __init__(self, parent):
        self.parent = parent
        self.ja_sacar = ctk.CTkToplevel(parent.root)
        self.ja_sacar.geometry('450x300')
        self.ja_sacar.title('Sacar')
        self.ja_sacar.maxsize(450, 300)
        self.ja_sacar.minsize(450, 300)
        self.ja_sacar.focus_force()
        self.ja_sacar.grab_set()

        self.initialize_ui()

    def initialize_ui(self):
        titulo = ctk.CTkLabel(self.ja_sacar, text='Sacar', text_color='#fff',
                              font=('Montserrat', 26, 'bold'))
        titulo.pack(pady=30)

        label_conta = ctk.CTkLabel(self.ja_sacar, text='Informe o numero da sua conta: ', text_color='#fff',
                                   font=('Montserrat', 17, 'bold'))
        label_conta.place(x=30, y=90)

        self.entry_conta = ctk.CTkEntry(self.ja_sacar, placeholder_text='Insira a conta', placeholder_text_color='#fff',
                                        font=('Montserrat', 12, 'bold'), fg_color='transparent')
        self.entry_conta.place(x=290, y=90)

        label_senha = ctk.CTkLabel(self.ja_sacar, text='Informe a senha: ', text_color='#fff',
                                   font=('Montserrat', 17, 'bold'))
        label_senha.place(x=30, y=130)

        self.entry_senha = ctk.CTkEntry(self.ja_sacar, placeholder_text='Insira a senha', placeholder_text_color='#fff',
                                        font=('Montserrat', 12, 'bold'), fg_color='transparent')
        self.entry_senha.place(x=170, y=130)

        label_valor = ctk.CTkLabel(self.ja_sacar, text='Informe o valor do saque: ', text_color='#fff',
                                   font=('Montserrat', 17, 'bold'))
        label_valor.place(x=30, y=170)

        self.entry_valor = ctk.CTkEntry(self.ja_sacar, placeholder_text='Insira o valor', placeholder_text_color='#fff',
                                        font=('Montserrat', 12, 'bold'), fg_color='transparent')
        self.entry_valor.place(x=240, y=170)

        sacar_button = ctk.CTkButton(self.ja_sacar, text='Fazer saque', text_color='#fff',
                                     font=('Montserrat', 18, 'bold'), command=self.saque)
        sacar_button.place(x=160, y=220)

    def saque(self):
        fichario = openpyxl.load_workbook('cadastros.xlsx')
        folha = fichario.active
        dados = []
        e_conta = self.entry_conta.get()
        e_valor = self.entry_valor.get()
        e_senha = self.entry_senha.get()

        for row in folha.iter_cols(values_only=True):
            dados.append(list(row))

        dicionario = {}

        for i in range(len(dados)):
            chave = dados[0][i]
            valores = [dados[j][i] for j in range(1, len(dados))]
            dicionario[chave] = valores

        if e_conta == "" or e_valor == "" or e_senha == "":
            tkinter.messagebox.showwarning('Sistema', "ERROR\nPor favor prenchar todos os campos!!")
        else:
            if int(e_conta) in dicionario:
                if dicionario[int(e_conta)][4] == str(e_senha):
                    if float(e_valor) != 0:
                        if float(e_valor) <= dicionario[int(e_conta)][5]:
                            for row in folha.iter_cols(min_row=2, max_row=folha.max_row, min_col=1, max_col=1):
                                for cell in row:
                                    if cell.value == int(e_conta):
                                        valor = cell.row
                                        folha[f'G{valor}'] = dicionario[int(e_conta)][5] - float(e_valor)
                            tkinter.messagebox.showwarning('Sistema', "Saque realizado com sucesso!!")
                            fichario.save(r'cadastros.xlsx')
                            self.entry_conta.delete(0, tkinter.END)
                            self.entry_senha.delete(0, tkinter.END)
                            self.entry_valor.delete(0, tkinter.END)
                        else:
                            tkinter.messagebox.showwarning('Sistema',
                                                           "ERROR\nVocê não tem saldo para realizar o saque!!")
                    else:
                        tkinter.messagebox.showwarning('Sistema', "ERROR\nValor invalido!!")
                else:
                    tkinter.messagebox.showwarning('Sistema', "ERROR\nSenha invalida!!")
            else:
                tkinter.messagebox.showwarning('Sistema', "ERROR\nConta invalida!!")


def main():
    root = ctk.CTk()
    app = BancoApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
